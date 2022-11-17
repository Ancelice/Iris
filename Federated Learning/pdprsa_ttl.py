import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
import ipaddress
from sklearn import metrics
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score
import openpyxl
from tensorflow.keras import layers, models


filename = os.listdir("C:/Users/anli/iris/FL/data")
print(filename)
print()

df = pd.read_excel("./data/" + filename[0], engine="openpyxl")
train_rsa = df['reply_src_addr']
train_pdp = df['probe_dst_prefix']
tmp1 = df['probe_ttl']
train_x = []
train_y = []
for z in range(len(tmp1)):
    train_x.append([int(ipaddress.IPv4Address(df.iat[z, 1])), int(ipaddress.IPv4Address(df.iat[z, 3]))])
    train_y.append(int(tmp1[z]))
train_x = np.array(train_x)
train_y = np.reshape(train_y, (-1))

MASTER = models.Sequential()
MASTER.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
MASTER.add(layers.Activation('relu'))
MASTER.add(layers.Dropout(0.5))
MASTER.add(layers.Dense(10))
MASTER.add(layers.Activation('softmax'))
MASTER.add(layers.Dense(1))
MASTER.compile(loss='mse',optimizer='Adam',metrics=['mae'])

for i in range(len(filename) - 1):

    MASMSE = []
    MASMAE = []
    MASTIME = []
    MASOBS = []
    MASPRED = []

    print()
    print()
    print('----- Epoch : ' + str(i + 1) + ' / ' + str(len(filename) - 1) + ' -----')
    print()

    # train
    df = pd.read_excel("./data/" + filename[i], engine="openpyxl")
    train_rsa = df['reply_src_addr']
    train_pdp = df['probe_dst_prefix']
    tmp1 = df['probe_ttl']
    train_x = []
    train_y = []
    for f in range(len(tmp1)):
        train_x.append([int(ipaddress.IPv4Address(df.iat[f, 1])), int(ipaddress.IPv4Address(df.iat[f, 3]))])
        train_y.append(int(tmp1[f]))
    train_x = np.array(train_x)
    train_y = np.reshape(train_y, (-1))
    print()
    print('=== MASTER / ' + filename[i] + ' ===')
    if len(train_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(train_x)) + ' / ' + str(train_x[0]))
        print('Train Y : ' + str(len(train_y)) + ' / ' + str(train_y[0]))
    print()

    #####

    # test
    df3 = pd.read_excel("./data/" + filename[i + 1], engine="openpyxl")
    test_rsa = df3['reply_src_addr']
    test_pdp = df3['reply_src_addr']
    tmp3 = df3['probe_ttl']
    MAStest_x = []
    MAStest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./data/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for g in range(len(tmp3)):
        MAStest_x.append([int(ipaddress.IPv4Address(df3.iat[g, 1])), int(ipaddress.IPv4Address(df3.iat[g, 3]))])
        MAStest_y.append(int(tmp3[g]))
        MASTIME.append(editName)
        MASOBS.append(int(tmp3[g]))
    MAStest_x = np.array(MAStest_x)
    MAStest_y = np.reshape(MAStest_y, (-1))
    print()
    print('=== MASTER / ' + filename[i + 1] + ' ===')
    if len(MAStest_x) == 0:
        print('Test X : 0')
        print('Test : 0')
    else :
        print('Test X : ' + str(len(MAStest_x)) + ' / ' + str(MAStest_x[0]))
        print('Test Y : ' + str(len(MAStest_y)) + ' / ' + str(MAStest_y[0]))
    print()

    #####

    # Learning and Predict
    MAShistory = MASTER.fit(train_x, train_y, epochs=1, verbose=1)

    LOSS, MAE = MASTER.evaluate(MAStest_x, MAStest_y, verbose=0)
    MASpred = MASTER.predict(MAStest_x)
    print('Master / pred : ' + str(MASpred[0]))
    print('Master / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for gg in range(len(MASpred)):
        MASPRED.append(MASpred[gg])
        MASMSE.append(LOSS)
        MASMAE.append(MAE)

    #####

    # Save file / Master
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for ww in header_list:
        ws[chr(x) + str(1)] = ww
        x = x + 1
    h = 2
    for yy in range(len(MASPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = MASMSE[yy]
        ws['C' + str(h) ] = MASMAE[yy]
        ws['D' + str(h) ] = MASOBS[yy]
        ws['E' + str(h) ] = int(MASPRED[yy])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/pdprsa_ttl/' +  tmpName + '.xlsx'
    print('Save Master : ' + savename)
    wb.save(savename)