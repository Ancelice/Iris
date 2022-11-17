import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
import ipaddress
from sklearn import metrics
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score
import openpyxl
from tensorflow.keras import layers, models


filename = os.listdir("C:/Users/anli/iris/GeoIP/reply_src_addr")
print(filename)
print()

df = pd.read_excel("./rsa/Asia/" + filename[0], engine="openpyxl")
train_rsa = df['reply_src_addr']
tmp1 = df['probe_ttl']
train_x = []
train_y = []
for z in range(len(tmp1)):
    train_x.append(int(ipaddress.IPv4Address(df.iat[z, 3])))
    train_y.append(int(tmp1[z]))
train_x = np.array(train_x).reshape(-1, 1)
train_y = np.reshape(train_y, (-1))

AF = models.Sequential()
AF.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
AF.add(layers.Activation('relu'))
AF.add(layers.Dropout(0.5))
AF.add(layers.Dense(10))
AF.add(layers.Activation('softmax'))
AF.add(layers.Dense(1))
AF.compile(loss='mse',optimizer='Adam',metrics=['mae'])

AS = models.Sequential()
AS.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
AS.add(layers.Activation('relu'))
AS.add(layers.Dropout(0.5))
AS.add(layers.Dense(10))
AS.add(layers.Activation('softmax'))
AS.add(layers.Dense(1))
AS.compile(loss='mse',optimizer='Adam',metrics=['mae'])

EU = models.Sequential()
EU.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
EU.add(layers.Activation('relu'))
EU.add(layers.Dropout(0.5))
EU.add(layers.Dense(10))
EU.add(layers.Activation('softmax'))
EU.add(layers.Dense(1))
EU.compile(loss='mse',optimizer='Adam',metrics=['mae'])

NA = models.Sequential()
NA.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
NA.add(layers.Activation('relu'))
NA.add(layers.Dropout(0.5))
NA.add(layers.Dense(10))
NA.add(layers.Activation('softmax'))
NA.add(layers.Dense(1))
NA.compile(loss='mse',optimizer='Adam',metrics=['mae'])

OC = models.Sequential()
OC.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
OC.add(layers.Activation('relu'))
OC.add(layers.Dropout(0.5))
OC.add(layers.Dense(10))
OC.add(layers.Activation('softmax'))
OC.add(layers.Dense(1))
OC.compile(loss='mse',optimizer='Adam',metrics=['mae'])

SA = models.Sequential()
SA.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
SA.add(layers.Activation('relu'))
SA.add(layers.Dropout(0.5))
SA.add(layers.Dense(10))
SA.add(layers.Activation('softmax'))
SA.add(layers.Dense(1))
SA.compile(loss='mse',optimizer='Adam',metrics=['mae'])

MASTER = models.Sequential()
MASTER.add(layers.Dense(64, input_shape = (train_x.shape[1], )))
MASTER.add(layers.Activation('relu'))
MASTER.add(layers.Dropout(0.5))
MASTER.add(layers.Dense(10))
MASTER.add(layers.Activation('softmax'))
MASTER.add(layers.Dense(1))
MASTER.compile(loss='mse',optimizer='Adam',metrics=['mae'])

for i in range(len(filename) - 1):

    AFMSE = []
    AFMAE = []
    AFTIME = []
    AFOBS = []
    AFPRED = []

    ASMSE = []
    ASMAE = []
    ASTIME = []
    ASOBS = []
    ASPRED = []

    EUMSE = []
    EUMAE = []
    EUTIME = []
    EUOBS = []
    EUPRED = []

    NAMSE = []
    NAMAE = []
    NATIME = []
    NAOBS = []
    NAPRED = []

    OCMSE = []
    OCMAE = []
    OCTIME = []
    OCOBS = []
    OCPRED = []

    SAMSE = []
    SAMAE = []
    SATIME = []
    SAOBS = []
    SAPRED = []

    MASMSE = []
    MASMAE = []
    MASTIME = []
    MASOBS = []
    MASPRED = []

    print()
    print()
    print('----- Epoch : ' + str(i + 1) + ' / ' + str(len(filename) - 1) + ' -----')

    # train / Africa
    df = pd.read_excel("./rsa/Africa/" + filename[i], engine="openpyxl")
    train_rsa = df['reply_src_addr']
    tmp1 = df['probe_ttl']
    AFtrain_x = []
    AFtrain_y = []
    for a in range(len(tmp1)):
        AFtrain_x.append(int(ipaddress.IPv4Address(df.iat[a, 3])))
        AFtrain_y.append(int(tmp1[a]))
    AFtrain_x = np.array(AFtrain_x).reshape(-1, 1)
    AFtrain_y = np.reshape(AFtrain_y, (-1))
    print()
    print('=== Africa / ' + filename[i] + ' ===')
    if len(AFtrain_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(AFtrain_x)) + ' / ' + str(AFtrain_x[0]))
        print('Train Y : ' + str(len(AFtrain_y)) + ' / ' + str(AFtrain_y[0]))
    print()  

    # train / Asia
    df = pd.read_excel("./rsa/Asia/" + filename[i], engine="openpyxl")
    train_rsa = ['reply_src_addr']
    tmp1 = df['probe_ttl']
    AStrain_x = []
    AStrain_y = []
    for b in range(len(tmp1)):
        AStrain_x.append(int(ipaddress.IPv4Address(df.iat[b, 3])))
        AStrain_y.append(int(tmp1[b]))
    AStrain_x = np.array(AStrain_x).reshape(-1, 1)
    AStrain_y = np.reshape(AStrain_y, (-1))
    print()
    print('=== Asia / ' + filename[i] + ' ===')
    if len(AStrain_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(AStrain_x)) + ' / ' + str(AStrain_x[0]))
        print('Train Y : ' + str(len(AStrain_y)) + ' / ' + str(AStrain_y[0]))
    print()   

    # train / Europe
    df = pd.read_excel("./rsa/Europe/" + filename[i], engine="openpyxl")
    train_rsa = df['reply_src_addr']
    tmp1 = df['probe_ttl']
    EUtrain_x = []
    EUtrain_y = []
    for c in range(len(tmp1)):
        EUtrain_x.append(int(ipaddress.IPv4Address(df.iat[c, 3])))
        EUtrain_y.append(int(tmp1[c]))
    EUtrain_x = np.array(EUtrain_x).reshape(-1, 1)
    EUtrain_y = np.reshape(EUtrain_y, (-1))
    print()
    print('=== Europe / ' + filename[i] + ' ===')
    if len(EUtrain_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(EUtrain_x)) + ' / ' + str(EUtrain_x[0]))
        print('Train Y : ' + str(len(EUtrain_y)) + ' / ' + str(EUtrain_y[0]))
    print()

    # train / NA
    df = pd.read_excel("./rsa/NA/" + filename[i], engine="openpyxl")
    train_rsa = df['reply_src_addr']
    tmp1 = df['probe_ttl']
    NAtrain_x = []
    NAtrain_y = []
    for d in range(len(tmp1)):
        NAtrain_x.append(int(ipaddress.IPv4Address(df.iat[d, 3])))
        NAtrain_y.append(int(tmp1[d]))
    NAtrain_x = np.array(NAtrain_x).reshape(-1, 1)
    NAtrain_y = np.reshape(NAtrain_y, (-1))
    print()
    print('=== NA / ' + filename[i] + ' ===')
    if len(NAtrain_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(NAtrain_x)) + ' / ' + str(NAtrain_x[0]))
        print('Train Y : ' + str(len(NAtrain_y)) + ' / ' + str(NAtrain_y[0]))
    print()

    # train / Oceania
    df = pd.read_excel("./rsa/Oceania/" + filename[i], engine="openpyxl")
    train_rsa = df['reply_src_addr']
    tmp1 = df['probe_ttl']
    OCtrain_x = []
    OCtrain_y = []
    for e in range(len(tmp1)):
        OCtrain_x.append(int(ipaddress.IPv4Address(df.iat[e, 3])))
        OCtrain_y.append(int(tmp1[e]))
    OCtrain_x = np.array(OCtrain_x).reshape(-1, 1)
    OCtrain_y = np.reshape(OCtrain_y, (-1))
    print()
    print('=== Oceania / ' + filename[i] + ' ===')
    if len(OCtrain_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(OCtrain_x)) + ' / ' + str(OCtrain_x[0]))
        print('Train Y : ' + str(len(OCtrain_y)) + ' / ' + str(OCtrain_y[0]))
    print()

    # train / SA
    df = pd.read_excel("./rsa/SA/" + filename[i], engine="openpyxl")
    train_rsa = df['reply_src_addr']
    tmp1 = df['probe_ttl']
    SAtrain_x = []
    SAtrain_y = []
    for f in range(len(tmp1)):
        SAtrain_x.append(int(ipaddress.IPv4Address(df.iat[f, 3])))
        SAtrain_y.append(int(tmp1[f]))
    SAtrain_x = np.array(SAtrain_x).reshape(-1, 1)
    SAtrain_y = np.reshape(SAtrain_y, (-1))
    print()
    print('=== SA / ' + filename[i] + ' ===')
    if len(SAtrain_x) == 0:
        print('Train X : 0')
        print('Train Y : 0')
    else :
        print('Train X : ' + str(len(SAtrain_x)) + ' / ' + str(SAtrain_x[0]))
        print('Train Y : ' + str(len(SAtrain_y)) + ' / ' + str(SAtrain_y[0]))
    print()

    #####

    # test / Africa
    df2 = pd.read_excel("./rsa/Africa/" + filename[i + 1], engine="openpyxl")
    test_rsa = df2['reply_src_addr']
    tmp2 = df2['probe_ttl']
    AFtest_x = []
    AFtest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./rsa/Africa/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for aa in range(len(tmp2)):
        AFtest_x.append(int(ipaddress.IPv4Address(df2.iat[aa, 3])))
        AFtest_y.append(int(tmp2[aa]))
        AFTIME.append(editName)
        AFOBS.append(int(tmp2[aa]))
    AFtest_x = np.array(AFtest_x).reshape(-1, 1)
    AFtest_y = np.reshape(AFtest_y, (-1))
    #print()
    #print('=== Africa / ' + filename[i + 1] + ' ===')
    #if len(AFtest_x) == 0:
    #    print('Test X : 0')
    #    print('Test Y : 0')
    #else:
    #    print('Test X : ' + str(len(AFtest_x)) + ' / ' + str(AFtest_x[0]))
    #    print('Test Y : ' + str(len(AFtest_y)) + ' / ' + str(AFtest_y[0]))
    #print() 

    # test / Asia
    df2 = pd.read_excel("./rsa/Asia/" + filename[i + 1], engine="openpyxl")
    test_rsa = df2['reply_src_addr']
    tmp2 = df2['probe_ttl']
    AStest_x = []
    AStest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./rsa/Asia/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for bb in range(len(tmp2)):
        AStest_x.append(int(ipaddress.IPv4Address(df2.iat[bb, 3])))
        AStest_y.append(int(tmp2[bb]))
        ASTIME.append(editName)
        ASOBS.append(int(tmp2[bb]))
    AStest_x = np.array(AStest_x).reshape(-1, 1)
    AStest_y = np.reshape(AStest_y, (-1))
    #print()
    #print('=== Asia / ' + filename[i + 1] + ' ===')
    #if len(AStest_x) == 0:
    #    print('Test X : 0')
    #    print('Test Y : 0')
    #else:
    #    print('Test X : ' + str(len(AStest_x)) + ' / ' + str(AStest_x[0]))
    #    print('Test Y : ' + str(len(AStest_y)) + ' / ' + str(AStest_y[0]))
    #print()   

    # test / Europe
    df2 = pd.read_excel("./rsa/Europe/" + filename[i + 1], engine="openpyxl")
    test_rsa = df2['reply_src_addr']
    tmp2 = df2['probe_ttl']
    EUtest_x = []
    EUtest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./rsa/Europe/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for cc in range(len(tmp2)):
        EUtest_x.append(int(ipaddress.IPv4Address(df2.iat[cc, 3])))
        EUtest_y.append(int(tmp2[cc]))
        EUTIME.append(editName)
        EUOBS.append(int(tmp2[cc]))
    EUtest_x = np.array(EUtest_x).reshape(-1, 1)
    EUtest_y = np.reshape(EUtest_y, (-1))
    #print()
    #print('=== Europe / ' + filename[i + 1] + ' ===')
    #if len(EUtest_x) == 0:
    #    print('Test X : 0')
    #    print('Test Y : 0')
    #else:
    #    print('Test X : ' + str(len(EUtest_x)) + ' / ' + str(EUtest_x[0]))
    #    print('Test Y : ' + str(len(EUtest_y)) + ' / ' + str(EUtest_y[0]))
    #print()  

    # test / NA
    df2 = pd.read_excel("./rsa/NA/" + filename[i + 1], engine="openpyxl")
    test_rsa = df2['reply_src_addr']
    tmp2 = df2['probe_ttl']
    NAtest_x = []
    NAtest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./rsa/NA/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for dd in range(len(tmp2)):
        NAtest_x.append(int(ipaddress.IPv4Address(df2.iat[dd, 3])))
        NAtest_y.append(int(tmp2[dd]))
        NATIME.append(editName)
        NAOBS.append(int(tmp2[dd]))
    NAtest_x = np.array(NAtest_x).reshape(-1, 1)
    NAtest_y = np.reshape(NAtest_y, (-1))
    #print()
    #print('=== NA / ' + filename[i + 1] + ' ===')
    #if len(NAtest_x) == 0:
    #    print('Test X : 0')
    #    print('Test Y : 0')
    #else:
    #    print('Test X : ' + str(len(NAtest_x)) + ' / ' + str(NAtest_x[0]))
    #    print('Test Y : ' + str(len(NAtest_y)) + ' / ' + str(NAtest_y[0]))
    #print() 

    # test / Oceania
    df2 = pd.read_excel("./rsa/Oceania/" + filename[i + 1], engine="openpyxl")
    test_rsa = df2['reply_src_addr']
    tmp2 = df2['probe_ttl']
    OCtest_x = []
    OCtest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./rsa/Oceania/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for ee in range(len(tmp2)):
        OCtest_x.append(int(ipaddress.IPv4Address(df2.iat[ee, 3])))
        OCtest_y.append(int(tmp2[ee]))
        OCTIME.append(editName)
        OCOBS.append(int(tmp2[ee]))
    OCtest_x = np.array(OCtest_x).reshape(-1, 1)
    OCtest_y = np.reshape(OCtest_y, (-1))
    #print()
    #print('=== Oceania / ' + filename[i + 1] + ' ===')
    #if len(OCtest_x) == 0:
    #    print('Test X : 0')
    #    print('Test Y : 0')
    #else:
    #    print('Test X : ' + str(len(OCtest_x)) + ' / ' + str(OCtest_x[0]))
    #    print('Test Y : ' + str(len(OCtest_y)) + ' / ' + str(OCtest_y[0]))
    #print() 

    # test / SA
    df2 = pd.read_excel("./rsa/SA/" + filename[i + 1], engine="openpyxl")
    test_rsa = df2['reply_src_addr']
    tmp2 = df2['probe_ttl']
    SAtest_x = []
    SAtest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./rsa/SA', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for ff in range(len(tmp2)):
        SAtest_x.append(int(ipaddress.IPv4Address(df2.iat[ff, 3])))
        SAtest_y.append(int(tmp2[ff]))
        SATIME.append(editName)
        SAOBS.append(int(tmp2[ff]))
    SAtest_x = np.array(SAtest_x).reshape(-1, 1)
    SAtest_y = np.reshape(SAtest_y, (-1))
    #print()
    #print('=== SA / ' + filename[i + 1] + ' ===')
    #if len(SAtest_x) == 0:
    #    print('Test X : 0')
    #    print('Test Y : 0')
    #else:
    #    print('Test X : ' + str(len(SAtest_x)) + ' / ' + str(SAtest_x[0]))
    #    print('Test Y : ' + str(len(SAtest_y)) + ' / ' + str(SAtest_y[0]))
    #print() 

    #####

    df3 = pd.read_excel("./data/log_" + filename[i + 1], engine="openpyxl")
    test_rsa = df3['reply_src_addr']
    tmp3 = df3['probe_ttl']
    MAStest_x = []
    MAStest_y = []
    editName = filename[i + 1]
    editName = editName.replace('./data/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for g in range(len(tmp3)):
        MAStest_x.append(int(ipaddress.IPv4Address(df3.iat[g, 3])))
        MAStest_y.append(int(tmp3[g]))
        MASTIME.append(editName)
        MASOBS.append(int(tmp3[g]))
    MAStest_x = np.array(MAStest_x).reshape(-1, 1)
    MAStest_y = np.reshape(MAStest_y, (-1))

    # Federated Learning
    print('- train -')
    if len(AFtrain_x) != 0:
        AFhistory = AF.fit(AFtrain_x, AFtrain_y, epochs=1, verbose=1)
    if len(AStrain_x) != 0:
        AShistory = AS.fit(AStrain_x, AStrain_y, epochs=1, verbose=1)
    if len(EUtrain_x) != 0:
        EUhistory = EU.fit(EUtrain_x, EUtrain_y, epochs=1, verbose=1)
    if len(NAtrain_x) != 0:
        NAhistory = NA.fit(NAtrain_x, NAtrain_y, epochs=1, verbose=1)
    if len(OCtrain_x) != 0:
        OChistory = OC.fit(OCtrain_x, OCtrain_y, epochs=1, verbose=1)
    if len(SAtrain_x) != 0:
        SAhistory = SA.fit(SAtrain_x, SAtrain_y, epochs=1, verbose=1)

    AFpred = AF.predict(MAStest_x)
    print('Africa / pred : ' + str(AFpred[0]))
    ASpred = AS.predict(MAStest_x)
    print('Asia/ pred : ' + str(ASpred[0]))
    EUpred = EU.predict(MAStest_x)
    print('Europe / pred : ' + str(EUpred[0]))
    NApred = NA.predict(MAStest_x)
    print('NA / pred : ' + str(NApred[0]))
    OCpred = OC.predict(MAStest_x)
    print('Oceania / pred : ' + str(OCpred[0]))
    SApred = SA.predict(MAStest_x)
    print('SA / pred : ' + str(SApred[0]))

    LOSS, MAE = AF.evaluate(MAStest_x, MAStest_y, verbose=0)
    print('Africa / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for aaa in range(len(AFpred)):
        AFPRED.append(AFpred[aaa])
        AFMSE.append(LOSS)
        AFMAE.append(MAE)
    
    LOSS, MAE = AS.evaluate(MAStest_x, MAStest_y, verbose=0)
    print('Asia / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for bbb in range(len(ASpred)):
        ASPRED.append(ASpred[bbb])
        ASMSE.append(LOSS)
        ASMAE.append(MAE)

    LOSS, MAE = EU.evaluate(MAStest_x, MAStest_y, verbose=0)
    print('Europe / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for ccc in range(len(EUpred)):
        EUPRED.append(EUpred[ccc])
        EUMSE.append(LOSS)
        EUMAE.append(MAE)

    LOSS, MAE = NA.evaluate(MAStest_x, MAStest_y, verbose=0)
    print('NA / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for ddd in range(len(NApred)):
        NAPRED.append(NApred[ddd])
        NAMSE.append(LOSS)
        NAMAE.append(MAE)

    LOSS, MAE = OC.evaluate(MAStest_x, MAStest_y, verbose=0)
    print('Oceania / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for eee in range(len(OCpred)):
        OCPRED.append(OCpred[eee])
        OCMSE.append(LOSS)
        OCMAE.append(MAE)

    LOSS, MAE = SA.evaluate(MAStest_x, MAStest_y, verbose=0)
    print('SA / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for fff in range(len(SApred)):
        SAPRED.append(SApred[fff])
        SAMSE.append(LOSS)
        SAMAE.append(MAE)

    # Federated Learning  / Server
    WAF = AF.get_weights()
    WAS = AS.get_weights()
    WEU = EU.get_weights()
    WNA = NA.get_weights()
    WOC = OC.get_weights()
    WSA = SA.get_weights()

    sumW = [(a + b + c + d + e + f) / 6 for (a, b, c, d, e, f) in zip(WAF, WAS, WEU, WNA, WOC, WSA)]

    MASTER.set_weights(sumW)

    LOSS, MAE = MASTER.evaluate(MAStest_x, MAStest_y, verbose=0)
    MASpred = MASTER.predict(MAStest_x)
    print('Master / pred : ' + str(MASpred[0]))
    print('Master / MSE : ' + str(LOSS) + ' / MAE : ' + str(MAE))
    for gg in range(len(MASpred)):
        MASPRED.append(MASpred[gg])
        MASMSE.append(LOSS)
        MASMAE.append(MAE)

    WMAS = MASTER.get_weights()
    AF.set_weights(WMAS)
    AS.set_weights(WMAS)
    EU.set_weights(WMAS)
    NA.set_weights(WMAS)
    OC.set_weights(WMAS)
    SA.set_weights(WMAS)

    #####


    # Save file / Africa
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for l in header_list:
        ws[chr(x) + str(1)] = l
        x = x + 1
    h = 2
    for m in range(len(AFPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = AFMSE[m]
        ws['C' + str(h) ] = AFMAE[m]
        ws['D' + str(h) ] = MASOBS[m]
        ws['E' + str(h) ] = int(AFPRED[m])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/Africa/' +  tmpName + '.xlsx'
    print('Save Africa : ' + savename)
    wb.save(savename)

    # Save file / Asia
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for o in header_list:
        ws[chr(x) + str(1)] = o
        x = x + 1
    h = 2
    for p in range(len(ASPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = ASMSE[p]
        ws['C' + str(h) ] = ASMAE[p]
        ws['D' + str(h) ] = MASOBS[p]
        ws['E' + str(h) ] = int(ASPRED[p])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/Asia/' +  tmpName + '.xlsx'
    print('Save Asia : ' + savename)
    wb.save(savename)

    # Save file / Europe
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for q in header_list:
        ws[chr(x) + str(1)] = q
        x = x + 1
    h = 2
    for r in range(len(EUPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = EUMSE[r]
        ws['C' + str(h) ] = EUMAE[r]
        ws['D' + str(h) ] = MASOBS[r]
        ws['E' + str(h) ] = int(EUPRED[r])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/Europe/' +  tmpName + '.xlsx'
    print('Save Europe : ' + savename)
    wb.save(savename)

    # Save file / NA
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for s in header_list:
        ws[chr(x) + str(1)] = s
        x = x + 1
    h = 2
    for t in range(len(NAPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = NAMSE[t]
        ws['C' + str(h) ] = NAMAE[t]
        ws['D' + str(h) ] = MASOBS[t]
        ws['E' + str(h) ] = int(NAPRED[t])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/NA/' +  tmpName + '.xlsx'
    print('Save NA : ' + savename)
    wb.save(savename)

    # Save file / Oceania
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for u in header_list:
        ws[chr(x) + str(1)] = u
        x = x + 1
    h = 2
    for v in range(len(OCPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = OCMSE[v]
        ws['C' + str(h) ] = OCMAE[v]
        ws['D' + str(h) ] = MASOBS[v]
        ws['E' + str(h) ] = int(OCPRED[v])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/Oceania/' +  tmpName + '.xlsx'
    print('Save Oceania : ' + savename)
    wb.save(savename)

    # Save file / SA
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for w in header_list:
        ws[chr(x) + str(1)] = w
        x = x + 1
    h = 2
    for y in range(len(SAPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = SAMSE[y]
        ws['C' + str(h) ] = SAMAE[y]
        ws['D' + str(h) ] = MASOBS[y]
        ws['E' + str(h) ] = int(SAPRED[y])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/SA/' +  tmpName + '.xlsx'
    print('Save SA : ' + savename)
    wb.save(savename)

    # Save file / Master
    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']
    header_list = ['Time', 'MSE', 'MAE', 'Observed', 'Predict']
    x = 65
    for ww in header_list:
        ws[chr(x) + str(1)] = ww
        x = x + 1
    h = 2
    for yy in range(len(MASPRED)):
        ws['A' + str(h) ] = MASTIME[0]
        ws['B' + str(h) ] = MASMSE[yy]
        ws['C' + str(h) ] = MASMAE[yy]
        ws['D' + str(h) ] = MASOBS[yy]
        ws['E' + str(h) ] = int(MASPRED[yy])
        h = h + 1
    tmpName = filename[i + 1]
    tmpName = editName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')
    savename = './result/rsa_ttl/Master/' +  tmpName + '.xlsx'
    print('Save Master : ' + savename)
    wb.save(savename)