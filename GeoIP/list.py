import requests
import json

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

import sys
import os
import plotly.express as px


filename = os.listdir("C:/Users/anli/iris/GeoIP/probe_dst_prefix")
print(filename)

df = pd.read_excel("./probe_dst_prefix/" + filename[len(filename)-1], engine="openpyxl")

for i in range(len(filename) - 1):
    df2  = pd.read_excel("./probe_dst_prefix/" + filename[len(filename) - i - 1], engine="openpyxl")
    df = pd.concat([df2, df], axis=0)

df = df[df['Frequency'] != 0]

print(df)

#####

import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['Time', 'Country', 'Frequency']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1

k = 2

for s in range (len(df)):
    ws['A' + str(k) ] = df.iat[s, 0]
    ws['B' + str(k) ] = df.iat[s, 1]
    ws['C' + str(k) ] = df.iat[s, 2]
    k = k + 1

wb.save('probe_dst_prefix.xlsx')