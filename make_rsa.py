import pandas as pd
import numpy as np
import sys
import os
import pycountry
import openpyxl
import requests
import json

filename = sys.argv[1]
print(filename)

#ISO code 2
ISO2 = ["AF", "AX", "AL", "DZ", "AS", "AD", "AO", "AI", "AQ", "AG", "AR",
"AM", "AW", "AU", "AT", "AZ", "BS", "BH", "BD", "BB", "BY", "BE",
"BZ", "BJ", "BM", "BT", "BO", "BQ", "BA", "BW", "BV", "BR", "IO",
"BN", "BG", "BF", "BI", "CV", "KH", "CM", "CA", "KY", "CF", "TD",
"CL", "CN", "CX", "CC", "CO", "KM", "CG", "CD", "CK", "CR", "CI",
"HR", "CU", "CW", "CY", "CZ", "DK", "DJ", "DM", "DO", "EC", "EG",
"SV", "GQ", "ER", "EE", "ET", "FK", "FO", "FJ", "FI", "FR", "GF",
"PF", "TF", "GA", "GM", "GE", "DE", "GH", "GI", "GR", "GL", "GD",
"GP", "GU", "GT", "GG", "GN", "GW", "GY", "HT", "HM", "VA", "HN",
"HK", "HU", "IS", "IN", "ID", "IR", "IQ", "IE", "IM", "IL", "IT",
"JM", "JP", "JE", "JO", "KZ", "KE", "KI", "KP", "KR", "KW", "KG",
"LA", "LV", "LB", "LS", "LR", "LY", "LI", "LT", "LU", "MO", "MK",
"MG", "MW", "MY", "MV", "ML", "MT", "MH", "MQ", "MR", "MU", "YT",
"MX", "FM", "MD", "MC", "MN", "ME", "MS", "MA", "MZ", "MM", "NA",
"NR", "NP", "NL", "NC", "NZ", "NI", "NE", "NG", "NU", "NF", "MP",
"NO", "OM", "PK", "PW", "PS", "PA", "PG", "PY", "PE", "PH", "PN",
"PL", "PT", "PR", "QA", "RE", "RO", "RU", "RW", "BL", "SH", "KN",
"LC", "MF", "PM", "VC", "WS", "SM", "ST", "SA", "SN", "RS", "SC",
"SL", "SG", "SX", "SK", "SI", "SB", "SO", "ZA", "GS", "SS", "ES",
"LK", "SD", "SR", "SJ", "SZ", "SE", "CH", "SY", "TW", "TJ", "TZ",
"TH", "TL", "TG", "TK", "TO", "TT", "TN", "TR", "TM", "TC", "TV",
"UG", "UA", "AE", "GB", "US", "UM", "UY", "UZ", "VU", "VE", "VN",
"VG", "VI", "WF", "EH", "YE", "ZM", "ZW"]

AF_psa = []
AF_pdp = []
AF_ttl = []
AF_rsa = []

AS_psa = []
AS_pdp = []
AS_ttl = []
AS_rsa = []

EU_psa = []
EU_pdp = []
EU_ttl = []
EU_rsa = []

NA_psa = []
NA_pdp = []
NA_ttl = []
NA_rsa = []

OC_psa = []
OC_pdp = []
OC_ttl = []
OC_rsa = []

SA_psa = []
SA_pdp = []
SA_ttl = []
SA_rsa = []

#####

# read Excel files

df = pd.read_excel(filename, engine="openpyxl")

data = pd.read_excel('./FL/list_rsa.xlsx', engine="openpyxl")
IP = []
Country = []
for z in range(len(data)):
    IP.append(data.iat[z, 0])
    Country.append(data.iat[z, 1])
data = ''

num = 0

while True:

    try:

        print()
        print(str(num) + " / " + str(len(df)))
    
        where = Country[IP.index(df.iat[num, 3])]

        # Africa
        if ( (where == 'GA') or 
             (where == 'DZ') or 
             (where == 'MA') or 
             (where == 'ZW') or 
             (where == 'RW') or 
             (where == 'MU') or 
             (where == 'BW') or 
             (where == 'ZA') ):

            AF_psa.append(df.iat[num, 0])
            AF_pdp.append(df.iat[num, 1])
            AF_ttl.append(df.iat[num, 2])
            AF_rsa.append(df.iat[num, 3])
            print('Africa')

        # Asia
        if ( (where == 'CN') or 
             (where == 'KR') or 
             (where == 'IL') or 
             (where == 'LA') or 
             (where == 'AE') or 
             (where == 'MO') or 
             (where == 'SA') or
             (where == 'JP') or
             (where == 'TH') or
             (where == 'SG') or
             (where == 'VN') or
             (where == 'AF') or
             (where == 'BD') or
             (where == 'HK') or 
             (where == 'IN') or 
             (where == 'MM') or 
             (where == 'NP') or 
             (where == 'PK') or 
             (where == 'PH') or 
             (where == 'TW') or 
             (where == 'TL') or 
             (where == 'KH') or 
             (where == 'BT') or 
             (where == 'IR') ):

            AS_psa.append(df.iat[num, 0])
            AS_pdp.append(df.iat[num, 1])
            AS_ttl.append(df.iat[num, 2])
            AS_rsa.append(df.iat[num, 3])
            print('Asia')

        # Europe
        if ( (where == 'DE') or 
             (where == 'FR') or 
             (where == 'BG') or 
             (where == 'HU') or 
             (where == 'AZ') or 
             (where == 'BY') or 
             (where == 'NO') or 
             (where == 'MD') or 
             (where == 'UA') or 
             (where == 'HR') or 
             (where == 'LI') or 
             (where == 'CZ') or 
             (where == 'DK') or 
             (where == 'RS') or 
             (where == 'FI') or 
             (where == 'RO') or 
             (where == 'TR') or 
             (where == 'BE') or 
             (where == 'KZ') or 
             (where == 'MC') or
             (where == 'PT') or
             (where == 'RU') or
             (where == 'GB') or
             (where == 'LT') or
             (where == 'LU') or
             (where == 'PL') or
             (where == 'IT') or
             (where == 'AT') or
             (where == 'GE') or 
             (where == 'IE') or 
             (where == 'NL') or 
             (where == 'SE') or 
             (where == 'ES') or 
             (where == 'CH') ):

            EU_psa.append(df.iat[num, 0])
            EU_pdp.append(df.iat[num, 1])
            EU_ttl.append(df.iat[num, 2])
            EU_rsa.append(df.iat[num, 3])
            print('Europe')

        # North America
        if ( (where == 'US') or 
             (where == 'CA') or 
             (where == 'PR') or 
             (where == 'FK') or
             (where == 'MX') ):

            NA_psa.append(df.iat[num, 0])
            NA_pdp.append(df.iat[num, 1])
            NA_ttl.append(df.iat[num, 2])
            NA_rsa.append(df.iat[num, 3])
            print('North America')

        # Oceania
        if ( (where == 'AU') or 
             (where == 'ID') or 
             (where == 'MP') or 
             (where == 'NZ') or
             (where == 'MY') or
             (where == 'AS') or
             (where == 'GU') ):

            OC_psa.append(df.iat[num, 0])
            OC_pdp.append(df.iat[num, 1])
            OC_ttl.append(df.iat[num, 2])
            OC_rsa.append(df.iat[num, 3])
            print('Oceania')

        # South America
        if ( (where == 'BR') or 
             (where == 'VE') or 
             (where == 'UY') or 
             (where == 'AR') or
             (where == 'GT') or 
             (where == 'CL') or
             (where == 'CO') or
             (where == 'EC') or
             (where == 'SV') or
             (where == 'NI') or
             (where == 'PA') or
             (where == 'PE') or 
             (where == 'CU') or 
             (where == 'EG') ):

            SA_psa.append(df.iat[num, 0])
            SA_pdp.append(df.iat[num, 1])
            SA_ttl.append(df.iat[num, 2])
            SA_rsa.append(df.iat[num, 3])
            print('South America')


    except json.decoder.JSONDecodeError as e:
        print (e)
        num = num - 1
    except:
        print ("Error !!")
        num = num - 1

    num = num + 1

    if num == len(df):
        break

#####

# Africa
wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']
header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
x = 65
for a in header_list:
    ws[chr(x) + str(1)] = a
    x = x + 1
k = 2
for b in range (len(AF_psa)):
    ws['A' + str(k) ] = AF_psa[b]
    ws['B' + str(k) ] = AF_pdp[b]
    ws['C' + str(k) ] = AF_ttl[b]
    ws['D' + str(k) ] = AF_rsa[b]
    k = k + 1
saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './FL/rsa/Africa/' + saveName
wb.save(saveName)
print('Africa : ' + str(len(AF_psa)))

# Asia
wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']
header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1
k = 2
for c in range (len(AS_psa)):
    ws['A' + str(k) ] = AS_psa[c]
    ws['B' + str(k) ] = AS_pdp[c]
    ws['C' + str(k) ] = AS_ttl[c]
    ws['D' + str(k) ] = AS_rsa[c]
    k = k + 1
saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './FL/rsa/Asia/' + saveName
wb.save(saveName)
print('Asia : ' + str(len(AS_psa)))

# Europe
wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']
header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
x = 65
for e in header_list:
    ws[chr(x) + str(1)] = e
    x = x + 1
k = 2
for f in range (len(EU_psa)):
    ws['A' + str(k) ] = EU_psa[f]
    ws['B' + str(k) ] = EU_pdp[f]
    ws['C' + str(k) ] = EU_ttl[f]
    ws['D' + str(k) ] = EU_rsa[f]
    k = k + 1
saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './FL/rsa/Europe/' + saveName
wb.save(saveName)
print('Europe : ' + str(len(EU_psa)))

# North America
wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']
header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
x = 65
for g in header_list:
    ws[chr(x) + str(1)] = g
    x = x + 1
k = 2
for h in range (len(NA_psa)):
    ws['A' + str(k) ] = NA_psa[h]
    ws['B' + str(k) ] = NA_pdp[h]
    ws['C' + str(k) ] = NA_ttl[h]
    ws['D' + str(k) ] = NA_rsa[h]
    k = k + 1
saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './FL/rsa/NA/' + saveName
wb.save(saveName)
print('North America : ' + str(len(NA_psa)))

# Oceania
wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']
header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
x = 65
for l in header_list:
    ws[chr(x) + str(1)] = l
    x = x + 1
k = 2
for m in range (len(OC_psa)):
    ws['A' + str(k) ] = OC_psa[m]
    ws['B' + str(k) ] = OC_pdp[m]
    ws['C' + str(k) ] = OC_ttl[m]
    ws['D' + str(k) ] = OC_rsa[m]
    k = k + 1
saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './FL/rsa/Oceania/' + saveName
wb.save(saveName)
print('Oceania : ' + str(len(OC_psa)))

# South America
wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']
header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
x = 65
for p in header_list:
    ws[chr(x) + str(1)] = p
    x = x + 1
k = 2
for q in range (len(SA_psa)):
    ws['A' + str(k) ] = SA_psa[q]
    ws['B' + str(k) ] = SA_pdp[q]
    ws['C' + str(k) ] = SA_ttl[q]
    ws['D' + str(k) ] = SA_rsa[q]
    k = k + 1
saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './FL/rsa/SA/' + saveName
wb.save(saveName)
print('South America : ' + str(len(SA_psa)))