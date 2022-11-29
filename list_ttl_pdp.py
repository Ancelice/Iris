import requests
import json
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import sys
import os
import plotly.express as px
import numpy as np
import openpyxl

filename = sys.argv[1]
print(filename)

#ISO code 2
ISO2 = ["AF", "AX", "AL", "DZ", "AS", "AD", "AO", "AI", "AQ", "AG", "AR",
"AM", "AW", "AU", "AT", "AZ", "BS", "BH", "BD", "BB", "BY", "BE",
"BZ", "BJ", "BM", "BT", "BO", "BQ", "BA", "BW", "BV", "BR", "IO",
"BN", "BG", "BF", "BI", "CV", "KH", "CM", "CA", "KY", "CF", "TD",
"CL", "CN", "CX", "CC", "CO", "KM", "CG", "CD", "CK", "CR", "CI",
"HR", "CU", "CW", "CY", "CZ", "DK", "DJ", "DM", "DO", "EC", "EG",
"SV", "GQ", "ER", "EE", "ET", "FK", "FO", "FJ", "FI", "FR", "GF",
"PF", "TF", "GA", "GM", "GE", "DE", "GH", "GI", "GR", "GL", "GD",
"GP", "GU", "GT", "GG", "GN", "GW", "GY", "HT", "HM", "VA", "HN",
"HK", "HU", "IS", "IN", "ID", "IR", "IQ", "IE", "IM", "IL", "IT",
"JM", "JP", "JE", "JO", "KZ", "KE", "KI", "KP", "KR", "KW", "KG",
"LA", "LV", "LB", "LS", "LR", "LY", "LI", "LT", "LU", "MO", "MK",
"MG", "MW", "MY", "MV", "ML", "MT", "MH", "MQ", "MR", "MU", "YT",
"MX", "FM", "MD", "MC", "MN", "ME", "MS", "MA", "MZ", "MM", "NA",
"NR", "NP", "NL", "NC", "NZ", "NI", "NE", "NG", "NU", "NF", "MP",
"NO", "OM", "PK", "PW", "PS", "PA", "PG", "PY", "PE", "PH", "PN",
"PL", "PT", "PR", "QA", "RE", "RO", "RU", "RW", "BL", "SH", "KN",
"LC", "MF", "PM", "VC", "WS", "SM", "ST", "SA", "SN", "RS", "SC",
"SL", "SG", "SX", "SK", "SI", "SB", "SO", "ZA", "GS", "SS", "ES",
"LK", "SD", "SR", "SJ", "SZ", "SE", "CH", "SY", "TW", "TJ", "TZ",
"TH", "TL", "TG", "TK", "TO", "TT", "TN", "TR", "TM", "TC", "TV",
"UG", "UA", "AE", "GB", "US", "UM", "UY", "UZ", "VU", "VE", "VN",
"VG", "VI", "WF", "EH", "YE", "ZM", "ZW"]

df = pd.read_excel(filename, engine="openpyxl")

TIME = []
editName = filename
editName = editName.replace('./data/', '')
editName = editName.replace('log_', '')
editName = editName.replace('.xlsx', '')
editName = editName[:4] + "/" + editName[4:]
editName = editName[:7] + "/" + editName[7:]
print(editName)

ttl = []
count = []
for i in range(len(ISO2)):
    count.append(0)
    ttl.append(0)
    TIME.append(editName)

#####

# read Excel files

data = pd.read_excel('./FL/list_pdp.xlsx', engine="openpyxl")
IP = []
Country = []
for z in range(len(data)):
    IP.append(data.iat[z, 0])
    Country.append(data.iat[z, 1])
data = ''

num = 0

error = 0

while True:

    try:

        print()
        print(str(num) + " / " + str(len(df)))

        where = Country[IP.index(df.iat[num, 1])]

        ttl[ISO2.index(where)] = ttl[ISO2.index(where)] + df.iat[num, 2]
        count[ISO2.index(where)] = count[ISO2.index(where)] + 1

    except json.decoder.JSONDecodeError as e:
        print (e)
        num = num - 1
        error = error + 1
    except:
        print ("Error !!")
        num = num - 1
        error = error + 1

    if (error == 10):
        num = num + 1
        error = 0

    num = num + 1

    if num == len(df):
        break


#####

import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['Time', 'Country', 'TTL']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1

k = 2

for yy in range(len(ttl)):
    if (ttl[yy] != 0):
        ttl[yy] = ttl[yy] / count[yy]

for s in range (len(TIME)):
    ws['A' + str(k) ] = TIME[s]
    ws['B' + str(k) ] = ISO2[s]
    ws['C' + str(k) ] = ttl[s]
    k = k + 1

saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './ttl_pdp/' + saveName
wb.save(saveName)