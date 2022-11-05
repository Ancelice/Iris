import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
import ipaddress
from sklearn import model_selection
from sklearn.ensemble import RandomForestRegressor
from sklearn import metrics
from sklearn.metrics import mean_squared_error
from sklearn.metrics import r2_score

#####

filename = os.listdir("C:/Users/anli/iris/data")
print(filename)
print()

df = pd.read_excel("./data/" + filename[0], engine="openpyxl")
train_rsa = df['probe_dst_prefix']
tmp1 = df['reply_src_addr']
train_x = []
train_y = []
for m in range(len(tmp1)):
    train_x.append(int(ipaddress.IPv4Address(df.iat[m, 1])))
    train_y.append(int(ipaddress.IPv4Address(df.iat[m, 3])))

#####

rf = RandomForestRegressor(random_state=0)

ACC = []
RMSE = []
TIME = []
OBS = []
PRED = []

for i in range(len(filename) - 1):

    print()
    print()
    print('----- Epoch : ' + str(i + 1) + ' -----')

    # train
    df = pd.read_excel("./data/" + filename[i], engine="openpyxl")
    train_rsa = df['probe_dst_prefix']
    tmp1 = df['reply_src_addr']
    train_x = []
    train_y = []
    for j in range(len(tmp1)):
        train_x.append(int(ipaddress.IPv4Address(df.iat[j, 1])))
        train_y.append(int(ipaddress.IPv4Address(df.iat[j, 3])))
    train_x = np.array(train_x).reshape(-1, 1)
    train_y = np.reshape(train_y, (-1))
    print()
    print('=== ' + filename[i] + ' ===')
    print('Train X : ' + str(len(train_x)) + ' / ' + str(train_x[0]))
    print('Train Y : ' + str(len(train_y)) + ' / ' + str(train_y[0]))
    print()    

    # test
    df2 = pd.read_excel("./data/" + filename[i + 1], engine="openpyxl")
    test_pdp = df2['probe_dst_prefix']
    tmp2 = df2['reply_src_addr']
    test_x = []
    test_y = []
    editName = filename[i + 1]
    editName = editName.replace('./data/', '')
    editName = editName.replace('log_', '')
    editName = editName.replace('.xlsx', '')
    editName = editName[:4] + "/" + editName[4:]
    editName = editName[:7] + "/" + editName[7:]
    for k in range(len(tmp2)):
        test_x.append(int(ipaddress.IPv4Address(df2.iat[k, 1])))
        test_y.append(int(ipaddress.IPv4Address(df2.iat[k, 3])))
        TIME.append(editName)
        OBS.append(test_pdp[k])
    test_x = np.array(test_x).reshape(-1, 1)
    test_y = np.reshape(test_y, (-1))
    print()
    print('=== ' + filename[i + 1] + ' ===')
    print('Test X : ' + str(len(test_x)) + ' / ' + str(test_x[0]))
    print('Test Y : ' + str(len(test_y)) + ' / ' + str(test_y[0]))
    print()  

    # Learning and Predict
    print('- train -')
    rf.fit(train_x,train_y)
    pred = rf.predict(test_x)
    print('pred : ' + str(pred[0]))
    print('- score -')
    score = r2_score(test_y, pred)
    rmse = np.sqrt(mean_squared_error(test_y, pred))
    for q in range(len(pred)):
        PRED.append(str(ipaddress.IPv4Address(int(pred[q]))))
        RMSE.append(rmse)
        ACC.append(score)
    print('R2-Score : ' + str(score))
    print('RMSE : ' + str(rmse))

#####

import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['Time', 'R2-Score', 'RMSE', 'Observed', 'Predict']
x = 65
for u in header_list:
    ws[chr(x) + str(1)] = u
    x = x + 1

h = 2

for p in data:
    ws['A' + str(h) ] = TIME[p]
    ws['B' + str(h) ] = ACC[p]
    ws['C' + str(h) ] = RMSE[p]
    ws['D' + str(h) ] = OBS[p]
    ws['D' + str(h) ] = PRED[p]
    h = h + 1

wb.save('./RESULT/pdp_rsa.xlsx')