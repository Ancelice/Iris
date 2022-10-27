import os
import schedule
from time import sleep
import datetime
from iris_client import IrisClient, AsyncIrisClient

def format_uuid(uuid):
    return uuid.replace("-", "_")

def get_tables(measurement, data_type="results"):
    """
    Get the table of each agent in the measurement.
    """
    return [
        f"{data_type}__{format_uuid(measurement['uuid'])}__{format_uuid(agent['agent_uuid'])}" 
        for agent in measurement["agents"]
    ]

def task():
    # Iris credentials
    base_url = "https://api.iris.dioptra.io"
    username = "mg22019@shibaura-it.ac.jp"
    password = "Hiro@114"
    #username = os.environ.get("IRIS_USERNAME")
    #password = os.environ.get("IRIS_PASSWORD")
    if not username or not password:
        raise ValueError("IRIS_USERNAME and IRIS_PASSWORD environment variables must be set")

    # Check out https://github.com/dioptra-io/iris-client for more info on IrisClient
    with IrisClient(base_url, username, password) as iris:
        measurements = iris.all("/measurements/public", params={"tag": "collection:exhaustive"})

    #####

    # We select the measurement we want to query (newest measurement)
    measurement = measurements[0]

    #####

    # Get temporary access token to the database for this measurement
    with IrisClient(base_url, username, password) as iris:
        me = iris.get("/users/me/services", params={"measurement_uuid": measurement["uuid"]}).json()
        credentials = dict(
            base_url=me["clickhouse"]["base_url"],
            database="iris",
            username=me["clickhouse"]["username"],
            password=me["clickhouse"]["password"]
        )

    #####

    def format_uuid(uuid):
        return uuid.replace("-", "_")

    def get_tables(measurement, data_type="results"):
        """
        Get the table of each agent in the measurement.
        """
        return [
            f"{data_type}__{format_uuid(measurement['uuid'])}__{format_uuid(agent['agent_uuid'])}" 
            for agent in measurement["agents"]
        ]

    #####

    # Get reply table names for this measurement
    tables = get_tables(measurement)

    #####

    data = []
    i = 0

    from pych_client import ClickHouseClient

    # Check out https://github.com/dioptra-io/pych-client for more info on ClickHouseClient
    with ClickHouseClient(**credentials) as client:
        for row in client.iter_json(
            f"SELECT probe_src_addr, probe_dst_prefix, probe_ttl, reply_src_addr FROM {tables[0]} LIMIT 100000"
        ):
            print(row)
            data.append(row)

            i = i + 1

    #####
    #print(type(row))

    #print(str(len(data[0])))
    #print(data[0])

    import openpyxl

    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']

    header_list = ['probe_src_addr', 'probe_dst_prefix', 'probe_ttl', 'reply_src_addr']
    x = 65
    for i in header_list:
        ws[chr(x) + str(1)] = i
        x = x + 1

    k = 2

    for i in data:
        ws['A' + str(k) ] = i['probe_src_addr'].replace('::ffff:', '')
        ws['B' + str(k) ] = i['probe_dst_prefix'].replace('::ffff:', '')
        ws['C' + str(k) ] = i['probe_ttl']
        ws['D' + str(k) ] = i['reply_src_addr'].replace('::ffff:', '')
        k = k + 1

    import datetime

    now = datetime.datetime.now()
    print(now)
    #filename = './data/log_' + now.strftime('%Y%m%d_%H%M%S') + '.xlsx'
    filename = './data/log_' + now.strftime('%Y%m%d_%H') + '.xlsx'
    wb.save(filename)

#####

schedule.every(1).hours.do(task)

while True:
    schedule.run_pending()
    sleep(1)