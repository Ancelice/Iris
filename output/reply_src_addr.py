import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
 
from sklearn.datasets import load_boston
 
import networkx as nx
from pyvis.network import Network

import sys

filename = sys.argv[1]
print(filename)

ind = []
val = []

df = pd.read_excel(filename, engine="openpyxl")
for index, value in df['reply_src_addr'].value_counts().iteritems():
    print(index, ': ', value)
    ind.append(str(index))
    val.append(value)

print(ind)
print(type(ind))
print(val)
print(type(val))

import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['reply_src_addr', 'frequency']
x = 65
for i in header_list:
    ws[chr(x) + str(1)] = i
    x = x + 1

k = 2

for i in range (len(ind)):
    ws['A' + str(k) ] = ind[i]
    ws['B' + str(k) ] = val[i]
    k = k + 1

wb.save('./frequency/frequency_100000.xlsx')

#vc.to_excel('./frequency/100.xlsx', index=False, header=False)