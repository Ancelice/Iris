import requests
import json

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

import sys
import os
import plotly.express as px

import pycountry

filename = sys.argv[1]
print(filename)

#####

#Probe_ttl
ind = []
#Frequency of occurrence
val = []

# read Excel files

df = pd.read_excel(filename, engine="openpyxl")
for index, value in df['probe_ttl'].value_counts().items():
    print(index, ': ', value)
    ind.append(str(index))
    val.append(value)

Time = []
editName = filename
editName = editName.replace('./data/', '')
editName = editName.replace('log_', '')
editName = editName.replace('.xlsx', '')
editName = editName[:4] + "/" + editName[4:]
editName = editName[:7] + "/" + editName[7:]
print(editName)
for h in range(len(ind)):
    Time.append(editName)

#####

import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['Time', 'TTL', 'Frequency']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1

k = 2

for s in range (len(ind)):
    ws['A' + str(k) ] = Time[s]
    ws['B' + str(k) ] = ind[s]
    ws['C' + str(k) ] = val[s]
    k = k + 1

saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './probe_ttl/' + saveName

wb.save(saveName)