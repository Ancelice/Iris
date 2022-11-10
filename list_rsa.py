import requests
import json

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

import sys
import os
import plotly.express as px

import openpyxl

#####

filename = os.listdir("C:/Users/anli/iris/data")
print(filename)

IP = []
Country = []

for i in range(len(filename)):

    ind = []
    val = []
    df  = pd.read_excel("./data/" + filename[i], engine="openpyxl")
    for index, value in df['reply_src_addr'].value_counts().items():
        print(index, ': ', value)
        ind.append(str(index))

    # Add IP and Country

    for j in range(len(ind)):

        num = 0
        print()
        print(filename[i] + " : " + str(j) + " / " + str(len(ind)))

        if (ind[j] not in IP):

            while True:

                try:

                    # IP address to test
                    ip_address = ind[j]
                    # URL to send the request to
                    request_url = 'https://geolocation-db.com/jsonp/' + ip_address
                    # Send request and decode the result
                    response = requests.get(request_url)
                    result = response.content.decode()
                    # Clean the returned string so it just contains the dictionary data for the IP address
                    result = result.split("(")[1].strip(")")
                    # Convert this data into a dictionary
                    result  = json.loads(result)
                    print(result)

                    IP.append(ind[j])
                    Country.append(result['country_code'])

                    num = 1

                except json.decoder.JSONDecodeError as e:
                    print (e)
                except:
                    print ("Error !!")

                if num == 1:
                    break

        else:
            ('Already in list.')


#####

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['reply_src_addr', 'Country']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1

k = 2

for s in range (len(IP)):
    ws['A' + str(k) ] = IP[s]
    ws['B' + str(k) ] = Country[s]
    k = k + 1

wb.save('./FL/list_rsa.xlsx')