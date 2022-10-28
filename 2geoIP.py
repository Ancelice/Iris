import requests
import json

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

import sys
import os
import plotly.express as px

import pycountry

filename = sys.argv[1]
print(filename)

#ISO code 2
ISO2 = ["AF", "AX", "AL", "DZ", "AS", "AD", "AO", "AI", "AQ", "AG", "AR",
"AM", "AW", "AU", "AT", "AZ", "BS", "BH", "BD", "BB", "BY", "BE",
"BZ", "BJ", "BM", "BT", "BO", "BQ", "BA", "BW", "BV", "BR", "IO",
"BN", "BG", "BF", "BI", "CV", "KH", "CM", "CA", "KY", "CF", "TD",
"CL", "CN", "CX", "CC", "CO", "KM", "CG", "CD", "CK", "CR", "CI",
"HR", "CU", "CW", "CY", "CZ", "DK", "DJ", "DM", "DO", "EC", "EG",
"SV", "GQ", "ER", "EE", "ET", "FK", "FO", "FJ", "FI", "FR", "GF",
"PF", "TF", "GA", "GM", "GE", "DE", "GH", "GI", "GR", "GL", "GD",
"GP", "GU", "GT", "GG", "GN", "GW", "GY", "HT", "HM", "VA", "HN",
"HK", "HU", "IS", "IN", "ID", "IR", "IQ", "IE", "IM", "IL", "IT",
"JM", "JP", "JE", "JO", "KZ", "KE", "KI", "KP", "KR", "KW", "KG",
"LA", "LV", "LB", "LS", "LR", "LY", "LI", "LT", "LU", "MO", "MK",
"MG", "MW", "MY", "MV", "ML", "MT", "MH", "MQ", "MR", "MU", "YT",
"MX", "FM", "MD", "MC", "MN", "ME", "MS", "MA", "MZ", "MM", "NA",
"NR", "NP", "NL", "NC", "NZ", "NI", "NE", "NG", "NU", "NF", "MP",
"NO", "OM", "PK", "PW", "PS", "PA", "PG", "PY", "PE", "PH", "PN",
"PL", "PT", "PR", "QA", "RE", "RO", "RU", "RW", "BL", "SH", "KN",
"LC", "MF", "PM", "VC", "WS", "SM", "ST", "SA", "SN", "RS", "SC",
"SL", "SG", "SX", "SK", "SI", "SB", "SO", "ZA", "GS", "SS", "ES",
"LK", "SD", "SR", "SJ", "SZ", "SE", "CH", "SY", "TW", "TJ", "TZ",
"TH", "TL", "TG", "TK", "TO", "TT", "TN", "TR", "TM", "TC", "TV",
"UG", "UA", "AE", "GB", "US", "UM", "UY", "UZ", "VU", "VE", "VN",
"VG", "VI", "WF", "EH", "YE", "ZM", "ZW"]

# ISO Code 3
ISO3 = []
for k in range(len(ISO2)):
    ISO3.append(pycountry.countries.get(alpha_2=ISO2[k]).alpha_3)

print(ISO3)
print(str(len(ISO3)))

iso_list = [0] * (len(ISO3))
print(iso_list)
print(str(len(iso_list)))

ISO = []
Time = []

editName = filename
editName = editName.replace('./data/', '')
editName = editName.replace('log_', '')
editName = editName.replace('.xlsx', '')
editName = editName[:4] + "/" + editName[4:]
editName = editName[:7] + "/" + editName[7:]
print(editName)
for h in range(len(ISO3)):
    ISO.append(ISO3[h])
    Time.append(editName)

#####

#Probe_dst_prefix
ind = []
#Frequency of occurrence
val = []
#Country Name ISO3
country = []

# read Excel files

df = pd.read_excel(filename, engine="openpyxl")
for index, value in df['reply_src_addr'].value_counts().items():
    print(index, ': ', value)
    ind.append(str(index))
    val.append(value)

num = 0

while True:

    try:

        print()
        print(str(num) + " / " + str(len(ind)))
    
        # IP address to test
        ip_address = ind[num]

        # URL to send the request to
        request_url = 'https://geolocation-db.com/jsonp/' + ip_address
        # Send request and decode the result
        response = requests.get(request_url)
        result = response.content.decode()
        # Clean the returned string so it just contains the dictionary data for the IP address
        result = result.split("(")[1].strip(")")
        # Convert this data into a dictionary
        result  = json.loads(result)
        print(result)
    
        #add frequency to iso_list
        count = 0
        for j in range(len(ISO2)):
            if (ISO2[count]) == (result['country_code']):
                print(count)
                iso_list[count] = iso_list[count] + val[num]
                print(pycountry.countries.get(alpha_2=result['country_code']).alpha_3)
                print(val[num])
            count = count + 1

    except json.decoder.JSONDecodeError as e:
        print (e)
        num = num - 1
    except:
        print ("Error !!")
        num = num - 1

    num = num + 1
    

    if num == len(ind):
        break

#####

import openpyxl

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['Time', 'Country', 'Frequency']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1

k = 2

for s in range (len(ISO)):
    ws['A' + str(k) ] = Time[s]
    ws['B' + str(k) ] = ISO[s]
    ws['C' + str(k) ] = iso_list[s]
    k = k + 1

saveName = filename
saveName = saveName.replace('./data/', '')
saveName = saveName.replace('log_', '')
saveName = './GeoIP/reply_src_addr/' + saveName

wb.save(saveName)