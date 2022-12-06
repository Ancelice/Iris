import requests
import json

import pandas as pd
import numpy as np
import sys
import os
import openpyxl

#####

filename = os.listdir("C:/Users/anli/iris/data")
print(filename)
print()

for i in range(len(filename)):

    pr = []
    ttl = []
    count = []

    df = pd.read_excel('./data/' + filename[i], engine="openpyxl")

    pr.append([df.iat[99999, 1], df.iat[99999, 3]])
    ttl.append(df.iat[99999, 2])
    count.append(1)

    for j in range(len(df) - 1):

        flag = -1

        if ([df.iat[j, 1], df.iat[j, 3]] not in pr):
            print(filename[i] + " : Add | " + str(j) + " / " + str(len(df) - 1))
            pr.append([df.iat[j, 1], df.iat[j, 3]])
            ttl.append(df.iat[j, 2])
            count.append(1)

        else:
            flag = pr.index([df.iat[j, 1], df.iat[j, 3]])
            count[flag] = count[flag] + 1

    print('List / ' + str(len(pr)))

    data = pd.DataFrame(pr, columns = ['pdp', 'rsa'])

    wb = openpyxl.load_workbook('result.xlsx')
    ws = wb['Sheet1']

    header_list = ['probe_dst_prefix', 'reply_src_addr', 'probe_ttl', 'count']
    x = 65
    for d in header_list:
        ws[chr(x) + str(1)] = d
        x = x + 1

    k = 2

    for s in range (len(data)):
        ws['A' + str(k) ] = data.iat[s, 0]
        ws['B' + str(k) ] = data.iat[s, 1]
        ws['C' + str(k) ] = ttl[s]
        ws['D' + str(k) ] = count[s]
        k = k + 1

    tmpName = filename[i]
    tmpName = tmpName.replace('./data', '')
    tmpName = tmpName.replace('log_', '')
    tmpName = tmpName.replace('/', '_')

    wb.save('./pdp_rsa/' + tmpName)