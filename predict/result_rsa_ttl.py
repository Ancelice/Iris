import requests
import json

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

import sys
import os
import plotly.express as px

import openpyxl

#####

filename = os.listdir("C:/Users/anli/iris/predict/rsa_ttl")
print(filename)

TIME = []
R2 = []
RMSE = []

for i in range(len(filename)):

    print(filename[i])
    df  = pd.read_excel("./rsa_ttl/" + filename[i], engine="openpyxl")
    TIME.append(i)
    R2.append(df.iat[1, 1])
    RMSE.append(df.iat[1, 2])

#####

wb = openpyxl.load_workbook('result.xlsx')
ws = wb['Sheet1']

header_list = ['Round', 'R2-Score', 'RMSE']
x = 65
for d in header_list:
    ws[chr(x) + str(1)] = d
    x = x + 1

k = 2

for s in range (len(TIME)):
    ws['A' + str(k) ] = TIME[s]
    ws['B' + str(k) ] = R2[s]
    ws['C' + str(k) ] = RMSE[s]
    k = k + 1

wb.save('Result_rsa_ttl.xlsx')

plt.plot(TIME, R2, label='R2-Score')
plt.ylim(-10, 10)
plt.title('R2-Score / rsa_ttl')
plt.show()

plt.plot(TIME, RMSE, label='RMSE')
plt.title('RMSE / rsa_ttl')
plt.show()